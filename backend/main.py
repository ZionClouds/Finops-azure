import os
import mimetypes
import requests
import smtplib
from email.message import EmailMessage
from datetime import datetime, timedelta
from typing import Dict, List
from openpyxl import Workbook
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from dotenv import load_dotenv
import time

load_dotenv()

app = FastAPI()

# ---------------------------------------------------
# Azure Authentication
# ---------------------------------------------------
def get_token(tenant_id, client_id, client_secret):
    url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    payload = {
        "client_id": client_id,
        "client_secret": client_secret,
        "grant_type": "client_credentials",
        "scope": "https://management.azure.com/.default"
    }
    response = requests.post(url, data=payload)
    response.raise_for_status()
    return response.json()["access_token"]

# ---------------------------------------------------
# Unified Cost Fetcher (Service + Resource Group level)
# ---------------------------------------------------
def fetch_combined_cost(subscription_id, token):
    url = f"https://management.azure.com/subscriptions/{subscription_id}/providers/Microsoft.CostManagement/query?api-version=2023-03-01"
    end_date = datetime.utcnow().date()
    start_date = end_date - timedelta(days=30)

    body = {
        "type": "Usage",
        "timeframe": "Custom",
        "timePeriod": {"from": str(start_date), "to": str(end_date)},
        "dataset": {
            "granularity": "None",
            "aggregation": {"totalCost": {"name": "PreTaxCost", "function": "Sum"}},
            "grouping": [
                {"type": "Dimension", "name": "ResourceGroupName"},
                {"type": "Dimension", "name": "ServiceName"}
            ]
        }
    }

    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    for attempt in range(5):
        response = requests.post(url, headers=headers, json=body)
        if response.status_code == 429:
            wait = int(response.headers.get("Retry-After", 10))
            time.sleep(wait)
            continue
        response.raise_for_status()
        break

    data = response.json()
    results = []
    for row in data["properties"]["rows"]:
        try:
            total_cost, resource_group, service, currency = row
            results.append({
                "resource_group": resource_group,
                "service": service,
                "total_cost": float(total_cost)
            })
        except Exception as e:
            print(f"⚠ Skipping row: {row} due to error: {e}")

    return results

# ---------------------------------------------------
# Excel Exporter
# ---------------------------------------------------
from collections import defaultdict

def generate_excel(all_costs: Dict[str, list], output_path: str):
    wb = Workbook()
    wb.remove(wb.active)
    grand_total = 0.0

    # Map subscription IDs to friendly names
    subscription_names = {
        "5c5c5028-4a7e-435c-8430-6ece5f592ae2": "zion-ai",
        "e21901bf-488a-4ded-b169-b694737e4c86": "zcs-admin"
    }

    for subscription_id, costs in all_costs.items():
        # Use friendly name or fallback to prefix
        friendly_name = subscription_names.get(subscription_id, subscription_id[:8])

        ### Sheet 1: Detailed breakdown (Resource Group, Service)
        sheet_detail = wb.create_sheet(title=f"{friendly_name}-Detail")
        sheet_detail.append(["Resource Group", "Service", "Total Cost"])
        subtotal = 0.0
        for entry in sorted(costs, key=lambda x: x["total_cost"], reverse=True):
            sheet_detail.append([
                entry["resource_group"],
                entry["service"],
                entry["total_cost"]
            ])
            subtotal += entry["total_cost"]
        sheet_detail.append([])
        sheet_detail.append(["Subtotal", "", subtotal])
        grand_total += subtotal

        ### Sheet 2: Aggregated by Resource Group
        sheet_rg = wb.create_sheet(title=f"{friendly_name}-RG-Summary")
        sheet_rg.append(["Resource Group", "Total Cost"])
        resource_group_totals = defaultdict(float)
        for entry in costs:
            resource_group_totals[entry["resource_group"]] += entry["total_cost"]
        for rg, total in sorted(resource_group_totals.items(), key=lambda x: x[1], reverse=True):
            sheet_rg.append([rg, total])

        ### Sheet 3: Aggregated by Service
        sheet_service = wb.create_sheet(title=f"{friendly_name}-Service-Summary")
        sheet_service.append(["Service Name", "Total Cost"])
        service_totals = defaultdict(float)
        for entry in costs:
            service_totals[entry["service"]] += entry["total_cost"]
        for service, total in sorted(service_totals.items(), key=lambda x: x[1], reverse=True):
            sheet_service.append([service, total])

    wb.save(output_path)
    return grand_total

# ---------------------------------------------------
# Email Sender (Dynamic Recipients)
# ---------------------------------------------------
def send_email(filepath, total_cost, recipients: List[str]):
    SMTP_SERVER = os.getenv("SMTP_SERVER")
    SMTP_PORT = int(os.getenv("SMTP_PORT"))
    EMAIL_USER = os.getenv("EMAIL_USER")
    EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")

    if not recipients:
        raise ValueError("Recipients list is empty")

    msg = EmailMessage()
    msg["Subject"] = "Azure Billing Summary"
    msg["From"] = EMAIL_USER
    msg["To"] = ", ".join(recipients)
    msg.set_content(
        f"Hello Team,\n\n"
        f"Please find attached the latest *Monthly Azure Billing Summary Report*.\n\n"
        f"Overall total cost across all resources (without discount): ${total_cost:.2f}\n\n"
        f"Regards,\n"
        f"DevOps Team"
    )

    file_type, _ = mimetypes.guess_type(filepath)
    with open(filepath, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype=file_type.split('/')[0],
            subtype=file_type.split('/')[1],
            filename=os.path.basename(filepath)
        )

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        server.send_message(msg)


# ---------------------------------------------------
# API Payload Model
# ---------------------------------------------------
class ScanRequest(BaseModel):
    recipients: List[str]

# ---------------------------------------------------
# Main API Endpoint
# ---------------------------------------------------
@app.post("/scan/azure-cost")
async def scan(request: ScanRequest):
    try:
        all_costs = {}
        index = 1
        while True:
            sub_id = os.getenv(f"SUBSCRIPTION_ID_{index}")
            if not sub_id:
                break

            tenant_id = os.getenv(f"TENANT_ID_{index}")
            client_id = os.getenv(f"CLIENT_ID_{index}")
            client_secret = os.getenv(f"CLIENT_SECRET_{index}")

            if not tenant_id or not client_id or not client_secret:
                raise ValueError(f"Missing credentials for subscription index {index}")

            token = get_token(tenant_id, client_id, client_secret)
            print(f"🔎 Scanning Subscription {sub_id}")

            costs = fetch_combined_cost(sub_id, token)
            all_costs[sub_id] = costs
            index += 1

        output_file = "/tmp/azure_cost_report.xlsx"
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        total_cost = generate_excel(all_costs, output_file)
        send_email(output_file, total_cost, request.recipients)

        return {"status": "success", "total_cost": total_cost}

    except Exception as e:
        print("🔥 Exception occurred:", str(e))
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8080, reload=True)