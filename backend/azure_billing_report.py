import os
import time
import requests
from collections import defaultdict
from urllib.parse import urlencode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

# --- Config ---
MONTHS = [
    ("2026-01-01", "2026-01-31", "Jan 2026"),
    ("2026-02-01", "2026-02-28", "Feb 2026"),
    ("2026-03-01", "2026-03-31", "Mar 2026"),
]
MONTH_LABELS = [m[2] for m in MONTHS]

SUBSCRIPTION_NAMES = {
    "5c5c5028-4a7e-435c-8430-6ece5f592ae2": "zion-ai",
    "e21901bf-488a-4ded-b169-b694737e4c86": "zcs-admin",
}

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "reports", "azure_billing_jan2026_mar2026.xlsx")

# --- Auth ---
def get_token(tenant_id, client_id, client_secret):
    url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    payload = {
        "client_id": client_id,
        "client_secret": client_secret,
        "grant_type": "client_credentials",
        "scope": "https://management.azure.com/.default",
    }
    response = requests.post(url, headers=headers, data=urlencode(payload))
    response.raise_for_status()
    print(f"  Token retrieved for tenant {tenant_id[:8]}...")
    return response.json()["access_token"]

# --- Cost Query ---
def fetch_monthly_cost(subscription_id, token, start_date, end_date):
    url = f"https://management.azure.com/subscriptions/{subscription_id}/providers/Microsoft.CostManagement/query?api-version=2023-03-01"
    body = {
        "type": "Usage",
        "timeframe": "Custom",
        "timePeriod": {"from": start_date, "to": end_date},
        "dataset": {
            "granularity": "None",
            "aggregation": {"totalCost": {"name": "PreTaxCost", "function": "Sum"}},
            "grouping": [
                {"type": "Dimension", "name": "ResourceGroupName"},
            ],
        },
    }
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    for attempt in range(5):
        response = requests.post(url, headers=headers, json=body)
        if response.status_code == 429:
            wait = int(response.headers.get("Retry-After", 10))
            print(f"    Rate limited, waiting {wait}s...")
            time.sleep(wait)
            continue
        response.raise_for_status()
        break

    data = response.json()
    results = {}
    for row in data.get("properties", {}).get("rows", []):
        try:
            total_cost, resource_group, currency = row
            results[resource_group] = round(float(total_cost), 2)
        except Exception as e:
            print(f"    Skipping row: {row} ({e})")
    return results

# --- Styles ---
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
total_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def apply_style(ws, row_num, col_count, font=None, fill=None):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.border = thin_border
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if col >= 2:
            cell.number_format = '#,##0.00'

# --- Main ---
def main():
    subscriptions = []
    index = 1
    while True:
        sub_id = os.getenv(f"SUBSCRIPTION_ID_{index}")
        if not sub_id:
            break
        subscriptions.append({
            "id": sub_id,
            "name": SUBSCRIPTION_NAMES.get(sub_id, f"sub-{index}"),
            "tenant_id": os.getenv(f"TENANT_ID_{index}"),
            "client_id": os.getenv(f"CLIENT_ID_{index}"),
            "client_secret": os.getenv(f"CLIENT_SECRET_{index}"),
        })
        index += 1

    print(f"Found {len(subscriptions)} subscription(s): {[s['name'] for s in subscriptions]}")

    # Fetch all data: all_data[sub_name][month_label] = {rg: cost}
    all_data = defaultdict(dict)

    for sub in subscriptions:
        print(f"\nSubscription: {sub['name']} ({sub['id']})")
        token = get_token(sub["tenant_id"], sub["client_id"], sub["client_secret"])

        for start_date, end_date, month_label in MONTHS:
            print(f"  Fetching {month_label}...")
            rg_costs = fetch_monthly_cost(sub["id"], token, start_date, end_date)
            all_data[sub["name"]][month_label] = rg_costs
            total = sum(rg_costs.values())
            print(f"    {month_label}: ${total:.2f} ({len(rg_costs)} resource groups)")

    # --- Build Excel ---
    wb = Workbook()
    wb.remove(wb.active)
    col_count = 2 + len(MONTH_LABELS)  # RG + months + Total

    for sub_name in all_data:
        ws = wb.create_sheet(title=sub_name)

        # Header: Resource Group | Jan 2026 | Feb 2026 | Mar 2026 | Total
        headers = ["Resource Group"] + [f"{m} ($)" for m in MONTH_LABELS] + ["Total ($)"]
        ws.append(headers)
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Collect all RGs across months
        all_rgs = set()
        for ml in MONTH_LABELS:
            all_rgs.update(all_data[sub_name].get(ml, {}).keys())

        # Sort RGs by total cost descending
        rg_totals = {}
        for rg in all_rgs:
            rg_totals[rg] = sum(
                all_data[sub_name].get(ml, {}).get(rg, 0) for ml in MONTH_LABELS
            )
        sorted_rgs = sorted(all_rgs, key=lambda x: rg_totals[x], reverse=True)

        # Data rows
        month_column_totals = defaultdict(float)
        for rg in sorted_rgs:
            row = [rg]
            rg_total = 0.0
            for ml in MONTH_LABELS:
                cost = all_data[sub_name].get(ml, {}).get(rg, 0)
                row.append(round(cost, 2))
                rg_total += cost
                month_column_totals[ml] += cost
            row.append(round(rg_total, 2))
            ws.append(row)
            apply_style(ws, ws.max_row, col_count)

        # Total row
        ws.append([])
        total_row = ["TOTAL"] + [round(month_column_totals[ml], 2) for ml in MONTH_LABELS] + [round(sum(month_column_totals.values()), 2)]
        ws.append(total_row)
        apply_style(ws, ws.max_row, col_count, total_font, total_fill)

        # Column widths
        ws.column_dimensions["A"].width = 45
        for i in range(2, col_count + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

        print(f"\n  Sheet '{sub_name}': {len(sorted_rgs)} resource groups, Total: ${sum(month_column_totals.values()):.2f}")

    # --- Save ---
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nReport saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
